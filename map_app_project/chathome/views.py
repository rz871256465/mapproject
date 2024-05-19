from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from map.models import Userinfo
from django.http import HttpResponse
from openpyxl import Workbook
from map.models import Userinfo
import json
import time
from game.models import Map
from django.contrib import messages
from .models import ChatMessage
from django.db import connection


def chathome(request):
    show_username = request.session.get("username", None)
    name_id = request.GET.get('num')
    map_objects = Map.objects.all()
    maplist = [map_obj.name for map_obj in map_objects]
    print(maplist)
    return render(request, 'chathome.html', {"name_id": name_id, "show_username": show_username, "map_objects": map_objects})


def chatpair(request, name_id):
    show_username = request.session.get("username", None)
    map_objects = Map.objects.all()
    maplist = [map_obj.name for map_obj in map_objects]
    print(maplist)
    return render(request, 'chathome.html', {"name_id": name_id, "show_username": show_username, "map_objects": map_objects})


def send_message(request):
    if request.method == 'POST':

        return HttpResponse("Message received successfully.", status=200)
    else:
        return HttpResponse("Only POST requests are allowed.", status=405)


def download_data_as_excel(request):
    # Query your SQL data using Django ORM
    users = Userinfo.objects.all()

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers to the first row
    headers = ['Username', 'Password', 'Email']
    worksheet.append(headers)

    # Write data rows
    for user in users:
        row = [user.user_name, user.user_pass, user.user_email]
        worksheet.append(row)

    # Create the HTTP response with Excel content type
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=user_data.xlsx'

    # Save the workbook content to the HTTP response
    workbook.save(response)

    return response


def save_maze(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        maze_data = data['maze']
        # 这里可以添加代码处理迷宫数据，例如保存到数据库

        return JsonResponse({'status': 'success', 'message': 'Maze data received and processed'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


roomnumber = 0
array = []


def timer1():
    time.sleep(1)


def checkArray(arrayList):
    count = 1
    while count <= 9:
        if (len(arrayList) > 1):
            count = 1
            return True
        else:
            count += 1
            timer1()
    return False
    # checkArray(arrayList)


def pair(request, otherUser, roomnumber):
    lists = []
    lists.append(roomnumber)
    print("User found")
    messages.success(
        request, f"You have been paired with ${otherUser}")
    show_username = request.session.get("username", None)
    name_id = roomnumber
    map_objects = Map.objects.all()
    maplist = [map_obj.name for map_obj in map_objects]
    print(maplist)
    return render(request, 'chathome.html', {"name_id": name_id, "show_username": show_username, "map_objects": map_objects})
    # return redirect(reverse('chathome', args=lists))


def pairing(request):
    if request.method == 'POST':
        global array
        global roomnumber
        userid = request.POST.get('userid')
        if len(array) == 0:
            array.append(userid)
            if (checkArray(array)):
                lists = []
                lists.append(roomnumber)
                print("User found")
                messages.success(
                    request, f"You have been paired with {array[1]}")
                show_username = request.session.get("username", None)
                name_id = roomnumber
                map_objects = Map.objects.all()
                maplist = [map_obj.name for map_obj in map_objects]
                print(maplist)
                # array = []
                return redirect('chatpair', name_id=name_id)
                # pair(request, array[1], roomnumber)
            else:
                array = []
                print("Print TimeOut")
                messages.error(
                    request, "No other Player Available. Try again in a few minutes.")
                print("No other Participant Available. Try again in a few minutes.")
                return redirect('map:indexpage')
        else:
            array.append(userid)
            time.sleep(8)
            # print("second")
            # user2 = array[0]
            # array = []
        # pass
        lists = []
        lists.append(roomnumber)
        print("User found")
        messages.success(
            request, f"You have been paired with {array[0]}")
        show_username = request.session.get("username", None)
        name_id = roomnumber
        map_objects = Map.objects.all()
        maplist = [map_obj.name for map_obj in map_objects]
        print(maplist)
        if roomnumber >= 9999999:
            roomnumber = 0
        roomnumber += 1
        array = []
        return redirect('chatpair', name_id=name_id)
        # return JsonResponse({'status': 'success', 'message': 'Maze data received and processed'})


def download_data_as_excel_another(request):
    # 连接到 MySQL 数据库
    username = request.session.get("username")
    table_name = f"chat_{username.lower()}"
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT id, message, created_at FROM {}".format(table_name))
        messages = cursor.fetchall()

    # 创建一个新的 Excel workbook
    workbook = Workbook()
    worksheet = workbook.active

    # 写入表头
    headers = ['Sender111111', 'Message', 'Timestamp']
    worksheet.append(headers)

    # 写入数据行
    for message in messages:
        worksheet.append(message)

    # 创建 HTTP 响应对象，设置内容类型和文件名
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=chat_messages.xlsx'

    # 保存 Excel 数据到 HTTP 响应对象
    workbook.save(response)

    # 关闭数据库连接
    cursor.close()
    connection.close()

    return response


def download_data_as_excel_map(request):
    # Query chat message data using Django ORM
    map_messages = Map.objects.all()

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers to the first row
    headers = ['name', 'maze']
    worksheet.append(headers)

    # Write data rows
    for map_messages in map_messages:
        maze_str = json.dumps(map_messages.maze)
        row = [map_messages.name, maze_str]
        worksheet.append(row)

    # Create the HTTP response with Excel content type
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=map_data.xlsx'

    # Save the workbook content to the HTTP response
    workbook.save(response)

    return response
